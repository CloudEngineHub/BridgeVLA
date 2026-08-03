#!/usr/bin/env python3
"""
Summarise RMBench evaluation results and rank them by success rate (an extension of summarize_rmbench_eval.py).

Validity: one evaluation (a given task / model_xxx / timestamp) counts only when it has all 100 episodes.
The success rate is computed straight from the per-episode detail lines (success / total).

Two kinds of output, written to the console, a .txt and several sheets of an .xlsx:

  1) Within-experiment ranking (PER_EXP_PREFIXES filters experiments by prefix; empty = all experiments)
     For each eval experiment, list every valid evaluation under it, best success rate first.
     Levels: experiment -> task -> model_xxx -> timestamp.

  2) Cross-experiment task ranking (scans every experiment)
     For each task, aggregate the valid 100-episode evaluations from all experiments, best first,
     annotating which experiment and model_xxx each row came from.

Directory layout:
  {LOG_ROOT}/{exp}/eval/rmbench/{task}/demo_clean/model_{N}/{timestamp}/_result.txt
  A few older experiments have no model level: {exp}/eval/rmbench/{task}/demo_clean/{timestamp}/_result.txt

Run:
  python summarize_rmbench_eval_ranked.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ======================== configuration ========================
# Root directory holding every experiment log
LOG_ROOT = Path(
    os.environ.get("BRIDGEVLA_LOG_DIR", "data/bridgevla_data/logs") + "/train_rmbench"
)

# The within-experiment ranking only covers experiments whose folder name starts with one of these prefixes;
# empty = all experiments.
PER_EXP_PREFIXES: Tuple[str, ...] = ()

# Only this split is counted
EVAL_SPLIT = "demo_clean"

# Minimum episode count for an evaluation to be valid
VALID_EPISODES = 100

# Output files (written under LOG_ROOT)
OUTPUT_XLSX = LOG_ROOT / "rmbench_eval_ranked.xlsx"
OUTPUT_TXT = LOG_ROOT / "rmbench_eval_ranked.txt"

MODEL_DIR_RE = re.compile(r"^model_")
NO_MODEL = "-"  # placeholder for old experiments that have no model_xxx level


@dataclass
class Run:
    """One evaluation result (a given experiment / task / model / timestamp)."""

    exp: str
    task: str
    model: str
    timestamp: str
    episodes: int
    success: int
    rate: float  # 0~1
    path: Path

    @property
    def is_valid(self) -> bool:
        return self.episodes == VALID_EPISODES


def model_sort_key(model_name: str) -> Tuple[int, str]:
    """Sort by checkpoint number, then lexicographically by suffix."""
    m = re.match(r"^model_(\d+)(?:_(.+))?$", model_name)
    if not m:
        return (10**9, model_name)
    return (int(m.group(1)), m.group(2) or "")


def parse_result(result_file: Path) -> Optional[Tuple[int, int]]:
    """Count (episodes, success) from the per-episode detail lines of _result.txt."""
    try:
        text = result_file.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return None

    episodes = success = 0
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("episode"):
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        episodes += 1
        if parts[1].strip().lower() == "success":
            success += 1
    return episodes, success


def iter_result_files(log_root: Path):
    """Locate _result.txt with two fixed-depth globs, avoiding a walk into the huge viz subtree.

    With a model level:    {exp}/eval/rmbench/{task}/{split}/{model}/{timestamp}/_result.txt
    Without a model level: {exp}/eval/rmbench/{task}/{split}/{timestamp}/_result.txt
    """
    seen = set()
    for pattern in (
        "*/eval/rmbench/*/*/*/*/_result.txt",
        "*/eval/rmbench/*/*/*/_result.txt",
    ):
        for f in log_root.glob(pattern):
            if f not in seen:
                seen.add(f)
                yield f


def collect_runs(log_root: Path) -> List[Run]:
    """Scan eval/rmbench under every experiment in LOG_ROOT and return all evaluation records."""
    runs: List[Run] = []

    for result_file in iter_result_files(log_root):
        parts = result_file.parts
        try:
            eval_idx = parts.index("eval")
            rmbench_idx = parts.index("rmbench", eval_idx)
        except ValueError:
            continue
        if eval_idx == 0:
            continue

        exp = parts[eval_idx - 1]
        try:
            task = parts[rmbench_idx + 1]
            split = parts[rmbench_idx + 2]
            after_split = parts[rmbench_idx + 3]
        except IndexError:
            continue

        if split != EVAL_SPLIT:
            continue

        # after_split may be model_xxx, or the timestamp directly (older experiments have no model level)
        if MODEL_DIR_RE.match(after_split):
            model = after_split
            try:
                timestamp = parts[rmbench_idx + 4]
            except IndexError:
                continue
        else:
            model = NO_MODEL
            timestamp = after_split

        parsed = parse_result(result_file)
        if parsed is None:
            continue
        episodes, success = parsed
        if episodes == 0:
            continue

        runs.append(
            Run(
                exp=exp,
                task=task,
                model=model,
                timestamp=timestamp,
                episodes=episodes,
                success=success,
                rate=success / episodes,
                path=result_file,
            )
        )

    return runs


def pct(rate: float) -> float:
    return round(rate * 100, 2)


# --------------------------- sort keys ---------------------------
def run_rank_key(run: Run) -> Tuple[float, Tuple[int, str], str]:
    """Success rate descending (negated), then model number, then timestamp."""
    return (-run.rate, model_sort_key(run.model), run.timestamp)


# --------------------------- console / TXT ---------------------------
def build_text_report(
    per_exp: Dict[str, List[Run]],
    by_task: Dict[str, List[Run]],
) -> str:
    lines: List[str] = []
    lines.append("=" * 78)
    lines.append(
        f"RMBench valid-evaluation ranking (valid = {VALID_EPISODES} episodes, split={EVAL_SPLIT})"
    )
    lines.append("=" * 78)

    # Output 1: within-experiment ranking
    lines.append("")
    lines.append("#" * 78)
    lines.append(f"# 1. Within-experiment ranking (prefixes: {', '.join(PER_EXP_PREFIXES) or 'all'})")
    lines.append("#" * 78)
    if not per_exp:
        lines.append("(no experiment matched the prefixes with a valid evaluation)")
    for exp in sorted(per_exp):
        exp_runs = sorted(per_exp[exp], key=run_rank_key)
        lines.append("")
        lines.append(f"=== experiment: {exp}  ({len(exp_runs)} valid evaluations) ===")
        lines.append(
            f"{'rank':<6}{'success':>9}  {'task':<22}{'model':<14}{'timestamp'}"
        )
        for i, r in enumerate(exp_runs, 1):
            lines.append(
                f"{i:<4}{pct(r.rate):>7.2f}%  {r.task:<22}{r.model:<14}{r.timestamp}"
            )

    # Output 2: cross-experiment ranking by task
    lines.append("")
    lines.append("#" * 78)
    lines.append("# 2. Cross-experiment task ranking (all experiments, grouped by task, best first)")
    lines.append("#" * 78)
    for task in sorted(by_task):
        task_runs = sorted(by_task[task], key=run_rank_key)
        lines.append("")
        lines.append(f"=== task: {task}  ({len(task_runs)} valid evaluations) ===")
        lines.append(
            f"{'rank':<6}{'success':>9}  {'experiment':<48}{'model':<14}{'timestamp'}"
        )
        for i, r in enumerate(task_runs, 1):
            lines.append(
                f"{i:<4}{pct(r.rate):>7.2f}%  {r.exp:<48}{r.model:<14}{r.timestamp}"
            )

    lines.append("")
    return "\n".join(lines)


# --------------------------- Excel ---------------------------
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
GROUP_FILL = PatternFill("solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def _style_header(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        c = ws.cell(row, col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = CENTER


def write_per_exp_sheet(ws, per_exp: Dict[str, List[Run]]) -> None:
    headers = ["experiment", "rank", "success (%)", "task", "model", "timestamp", "success/total"]
    ncol = len(headers)
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header(ws, 1, ncol)

    row = 2
    for exp in sorted(per_exp):
        exp_runs = sorted(per_exp[exp], key=run_rank_key)
        for i, r in enumerate(exp_runs, 1):
            values = [
                exp if i == 1 else "",
                i,
                pct(r.rate),
                r.task,
                r.model,
                r.timestamp,
                f"{r.success}/{r.episodes}",
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row, col, v)
                c.border = BORDER
                c.alignment = LEFT if col in (1, 4, 6) else CENTER
                if col == 3:
                    c.number_format = "0.00"
                if col == 1 and v:
                    c.font = Font(bold=True)
                    c.fill = GROUP_FILL
            row += 1

    _set_widths(ws, [40, 6, 11, 22, 14, 40, 12])


def write_by_task_sheet(ws, by_task: Dict[str, List[Run]]) -> None:
    headers = ["task", "rank", "success (%)", "experiment", "model", "timestamp", "success/total"]
    ncol = len(headers)
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header(ws, 1, ncol)

    row = 2
    for task in sorted(by_task):
        task_runs = sorted(by_task[task], key=run_rank_key)
        for i, r in enumerate(task_runs, 1):
            values = [
                task if i == 1 else "",
                i,
                pct(r.rate),
                r.exp,
                r.model,
                r.timestamp,
                f"{r.success}/{r.episodes}",
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row, col, v)
                c.border = BORDER
                c.alignment = LEFT if col in (1, 4, 6) else CENTER
                if col == 3:
                    c.number_format = "0.00"
                if col == 1 and v:
                    c.font = Font(bold=True)
                    c.fill = GROUP_FILL
            row += 1

    _set_widths(ws, [22, 6, 11, 44, 14, 40, 12])


def _set_widths(ws, widths: List[int]) -> None:
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_excel(
    per_exp: Dict[str, List[Run]],
    by_task: Dict[str, List[Run]],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "within-experiment"
    write_per_exp_sheet(ws1, per_exp)

    ws2 = wb.create_sheet("cross-experiment by task")
    write_by_task_sheet(ws2, by_task)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    log_root = LOG_ROOT.expanduser().resolve()
    if not log_root.is_dir():
        raise FileNotFoundError(f"log root does not exist: {log_root}")

    runs = collect_runs(log_root)
    valid_runs = [r for r in runs if r.is_valid]

    # Output 1: within-experiment ranking (selected prefixes only)
    per_exp: Dict[str, List[Run]] = {}
    for r in valid_runs:
        if not PER_EXP_PREFIXES or r.exp.startswith(PER_EXP_PREFIXES):
            per_exp.setdefault(r.exp, []).append(r)

    # Output 2: cross-experiment ranking by task (all experiments)
    by_task: Dict[str, List[Run]] = {}
    for r in valid_runs:
        by_task.setdefault(r.task, []).append(r)

    report = build_text_report(per_exp, by_task)
    print(report)

    OUTPUT_TXT.write_text(report, encoding="utf-8")
    write_excel(per_exp, by_task, OUTPUT_XLSX)

    total = len(runs)
    print("-" * 78)
    print(
        f"Scanned {total} result files, {len(valid_runs)} of them valid ({VALID_EPISODES} episodes)."
    )
    print(f"Within-experiment ranking covers {len(per_exp)} experiments; {len(by_task)} tasks across experiments.")
    print(f"Wrote: {OUTPUT_TXT}")
    print(f"Wrote: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
