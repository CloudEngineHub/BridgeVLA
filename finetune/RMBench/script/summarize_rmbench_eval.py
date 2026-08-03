#!/usr/bin/env python3
"""
Summarise the success rate per task / model from the RMBench evaluation logs and produce an Excel sheet.

Directory layout:
  {EXP_PATH}/eval/rmbench/{task_dir}/demo_clean/model_{N}/{timestamp}/_result.txt

Run:
  python summarize_rmbench_eval.py
"""

from __future__ import annotations

import re
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ======================== configuration ========================
# The ckpt directory to summarise: eval results land under its eval/rmbench/<task>/<cfg>/...
# (see eval_policy_client.py). Released ckpt directories and your own training runs share the same layout,
# so both can be summarised directly; point RMBENCH_SUMMARY_DIR at one instead of editing this file.
EXP_PATH = Path(
    os.environ.get("RMBENCH_SUMMARY_DIR")
    or (os.environ.get("BRIDGEVLA_RELEASE_CKPT_DIR", "data/bridgevla_ckpt/bridgevla_plus") + "/rmbench")
)

# Output filename prefix (written under EXP_PATH as {prefix}_{experiment}.xlsx)
OUTPUT_FILENAME_PREFIX = "rmbench_eval_summary"

# Only results under demo_clean are counted
EVAL_SPLIT = "demo_clean"

# (directory name, display name, TMC marker)
TASK_ROWS: List[Tuple[str, str, str]] = [
    ("observe_and_pickup", "Observe and Pick Up", "M(1)"),
    ("rearrange_blocks", "Rearrange Blocks", "M(1)"),
    ("put_back_block", "Put Back Block", "M(1)"),
    ("swap_blocks", "Swap Blocks", "M(1)"),
    ("swap_T", "Swap T", "M(1)"),
    ("battery_try", "Battery Try", "M(n)"),
    ("blocks_ranking_try", "Blocks Ranking Try", "M(n)"),
    ("cover_blocks", "Cover Blocks", "M(n)"),
    ("press_button", "Press Button", "M(n)"),
]

GROUP1_SIZE = 5
MODEL_DIR_RE = re.compile(r"^model_(\d+)(?:_(.+))?$")


def model_sort_key(model_name: str) -> Tuple[int, str]:
    """Sort by checkpoint number, then lexicographically by suffix."""
    match = MODEL_DIR_RE.match(model_name)
    if not match:
        return (999999, model_name)
    return (int(match.group(1)), match.group(2) or "")


def parse_success_rate(result_file: Path) -> Optional[float]:
    """Parse the success rate (a 0..1 fraction) out of _result.txt."""
    try:
        text = result_file.read_text(encoding="utf-8")
    except OSError:
        return None

    lines = [line.strip() for line in text.splitlines()]
    in_rate_block = False
    for line in lines:
        if line.startswith("Instruction Type:"):
            in_rate_block = True
            continue
        if not in_rate_block:
            continue
        if not line:
            continue
        if line.startswith("===="):
            break
        try:
            return float(line)
        except ValueError:
            continue

    # Fallback: count from the per-episode details
    success = total = 0
    for line in lines:
        if not line.startswith("episode"):
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        total += 1
        if parts[1].strip().lower() == "success":
            success += 1
    if total == 0:
        return None
    return success / total


def collect_results(exp_path: Path) -> Tuple[Dict[str, Dict[str, float]], List[str]]:
    """
    Scan the eval directory and return:
      results[task_dir][model_name] = success_rate (0~1)
      model_names: the sorted list of model_xx names
    """
    eval_root = exp_path / "eval" / "rmbench"
    if not eval_root.is_dir():
        raise FileNotFoundError(f"evaluation directory not found: {eval_root}")

    # (task, model_name) -> (timestamp, rate)
    latest: Dict[Tuple[str, str], Tuple[str, float]] = {}

    for result_file in eval_root.rglob("_result.txt"):
        parts = result_file.parts
        try:
            rmbench_idx = parts.index("rmbench")
            task_dir = parts[rmbench_idx + 1]
            split = parts[rmbench_idx + 2]
            model_dir = parts[rmbench_idx + 3]
            timestamp = parts[rmbench_idx + 4]
        except (ValueError, IndexError):
            continue

        if split != EVAL_SPLIT:
            continue
        if not MODEL_DIR_RE.match(model_dir):
            continue

        rate = parse_success_rate(result_file)
        if rate is None:
            continue

        key = (task_dir, model_dir)
        if key not in latest or timestamp > latest[key][0]:
            latest[key] = (timestamp, rate)

    results: Dict[str, Dict[str, float]] = {}
    model_names = sorted({model_name for _, model_name in latest}, key=model_sort_key)

    for (task_dir, model_name), (_, rate) in latest.items():
        results.setdefault(task_dir, {})[model_name] = rate

    return results, model_names


def mean(values: List[Optional[float]]) -> Optional[float]:
    valid = [v for v in values if v is not None]
    if not valid:
        return None
    return sum(valid) / len(valid)


def format_percent(rate: Optional[float]) -> Optional[float]:
    """Convert a 0..1 success rate to a percentage with two decimals (no % sign)."""
    if rate is None:
        return None
    return round(rate * 100, 2)


def build_table_rows(
    results: Dict[str, Dict[str, float]], model_names: List[str]
) -> List[Tuple[str, str, List[Optional[float]]]]:
    """Build the data rows: (task_label, tmc, [rates...]), including Average / Total Average."""
    rows: List[Tuple[str, str, List[Optional[float]]]] = []

    group1_rates: List[List[Optional[float]]] = []
    for task_dir, label, tmc in TASK_ROWS[:GROUP1_SIZE]:
        task_rates = [results.get(task_dir, {}).get(m) for m in model_names]
        rows.append((label, tmc, task_rates))
        group1_rates.append(task_rates)

    rows.append(
        (
            "Average",
            "M(1)",
            [mean([r[i] for r in group1_rates]) for i in range(len(model_names))],
        )
    )

    group2_rates: List[List[Optional[float]]] = []
    for task_dir, label, tmc in TASK_ROWS[GROUP1_SIZE:]:
        task_rates = [results.get(task_dir, {}).get(m) for m in model_names]
        rows.append((label, tmc, task_rates))
        group2_rates.append(task_rates)

    rows.append(
        (
            "Average",
            "M(n)",
            [mean([r[i] for r in group2_rates]) for i in range(len(model_names))],
        )
    )

    all_rates = group1_rates + group2_rates
    rows.append(
        (
            "Total Average",
            "/",
            [mean([r[i] for r in all_rates]) for i in range(len(model_names))],
        )
    )
    return rows


def write_excel(
    exp_path: Path,
    exp_name: str,
    model_names: List[str],
    table_rows: List[Tuple[str, str, List[Optional[float]]]],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RMBench Eval"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    avg_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # Column layout: A=Tasks, B=TMC, C...=models
    first_model_col = 3
    last_model_col = first_model_col + len(model_names) - 1

    ws.cell(1, 1, "Tasks")
    ws.cell(1, 2, "TMC")
    if model_names:
        ws.merge_cells(
            start_row=1,
            start_column=first_model_col,
            end_row=1,
            end_column=last_model_col,
        )
        title_cell = ws.cell(1, first_model_col, exp_name)
        title_cell.alignment = center
        title_cell.font = Font(bold=True)

        for idx, model_name in enumerate(model_names):
            cell = ws.cell(2, first_model_col + idx, model_name)
            cell.alignment = center
            cell.font = Font(bold=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(1, 1).font = Font(bold=True)
    ws.cell(1, 2).font = Font(bold=True)
    ws.cell(1, 1).alignment = center
    ws.cell(1, 2).alignment = center

    for row_idx, (label, tmc, rates) in enumerate(table_rows, start=3):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, tmc)
        is_avg_row = label in {"Average", "Total Average"}
        for col_offset, rate in enumerate(rates):
            value = format_percent(rate)
            cell = ws.cell(row_idx, first_model_col + col_offset, value)
            if value is not None:
                cell.number_format = "0.00"

        for col in range(1, last_model_col + 1):
            cell = ws.cell(row_idx, col)
            cell.border = border
            cell.alignment = left if col == 1 else center
            if is_avg_row:
                cell.fill = avg_fill
                cell.font = Font(bold=True)

    for col in range(1, last_model_col + 1):
        for row in (1, 2):
            cell = ws.cell(row, col)
            cell.border = border
            cell.fill = header_fill
            if row == 2 and col >= first_model_col:
                cell.alignment = center

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    for col in range(first_model_col, last_model_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    exp_path = EXP_PATH.expanduser().resolve()
    if not exp_path.is_dir():
        raise FileNotFoundError(f"experiment path does not exist: {exp_path}")

    exp_name = exp_path.name
    results, model_names = collect_results(exp_path)
    if not model_names:
        raise RuntimeError(
            f"no model_xx evaluation results found under {exp_path / 'eval' / 'rmbench'}"
        )

    table_rows = build_table_rows(results, model_names)
    output_path = exp_path / f"{OUTPUT_FILENAME_PREFIX}_{exp_name}.xlsx"
    write_excel(exp_path, exp_name, model_names, table_rows, output_path)

    print(f"Experiment: {exp_name}")
    print(f"Models: {', '.join(model_names)}")
    print(f"Wrote: {output_path}")


if __name__ == "__main__":
    main()
