#!/usr/bin/env python3
"""
Summarise the evaluation results of the RMBench no_spatial single-task experiments and rank them by success rate.

Adapted from summarize_rmbench_eval_ranked.py:
  - only scans single-task experiments whose name contains no_spatial;
  - parses the task out of the experiment name (new_rot_{task}_no_spatial...);
  - keeps only records whose eval-directory task matches the experiment-name task;
  - valid sample: an evaluation with all 100 episodes.

Output (console + .txt + .xlsx):
  1) within-experiment ranking: valid evaluations under each no_spatial experiment, best success rate first;
  2) cross-experiment task ranking: valid evaluations of all no_spatial experiments, grouped by task.

Directory layout:
  {LOG_ROOT}/{exp}/eval/rmbench/{task}/demo_clean/model_{N}/{timestamp}/_result.txt

Run:
  python summarize_rmbench_eval_no_spatial.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ======================== configuration ========================
LOG_ROOT = Path(
    os.environ.get("BRIDGEVLA_LOG_DIR", "data/bridgevla_data/logs") + "/train_rmbench"
)

EVAL_SPLIT = "demo_clean"
VALID_EPISODES = 100

OUTPUT_XLSX = LOG_ROOT / "rmbench_eval_no_spatial_ranked.xlsx"
OUTPUT_TXT = LOG_ROOT / "rmbench_eval_no_spatial_ranked.txt"

MODEL_DIR_RE = re.compile(r"^model_")
NO_MODEL = "-"
NO_SPATIAL_MARKER = "no_spatial"
EXP_TASK_RE = re.compile(r"new_rot_(.+?)_no_spatial")

KNOWN_TASKS: Tuple[str, ...] = (
    "blocks_ranking_try",
    "observe_and_pickup",
    "rearrange_blocks",
    "put_back_block",
    "cover_blocks",
    "press_button",
    "swap_blocks",
    "battery_try",
    "swap_T",
)


@dataclass
class Run:
    exp: str
    exp_task: str
    task: str
    model: str
    timestamp: str
    episodes: int
    success: int
    rate: float
    path: Path

    @property
    def is_valid(self) -> bool:
        return self.episodes == VALID_EPISODES


def extract_task_from_exp(exp: str) -> Optional[str]:
    """Parse the task out of a no_spatial single-task experiment name."""
    if NO_SPATIAL_MARKER not in exp:
        return None
    m = EXP_TASK_RE.search(exp)
    if m:
        return m.group(1)
    for task in sorted(KNOWN_TASKS, key=len, reverse=True):
        needle = f"_{task}_"
        if needle in exp:
            return task
    return None


def model_sort_key(model_name: str) -> Tuple[int, str]:
    m = re.match(r"^model_(\d+)(?:_(.+))?$", model_name)
    if not m:
        return (10**9, model_name)
    return (int(m.group(1)), m.group(2) or "")


def parse_result(result_file: Path) -> Optional[Tuple[int, int]]:
    try:
        text = result_file.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return None

    episodes = success = 0
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("episode"):
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        episodes += 1
        if parts[1].strip().lower() == "success":
            success += 1
    return episodes, success


def iter_result_files(log_root: Path):
    seen = set()
    for pattern in (
        "*/eval/rmbench/*/*/*/*/_result.txt",
        "*/eval/rmbench/*/*/*/_result.txt",
    ):
        for f in log_root.glob(pattern):
            if f not in seen:
                seen.add(f)
                yield f


def collect_runs(log_root: Path) -> Tuple[List[Run], List[str]]:
    """Scan the no_spatial single-task experiments and return (runs, skipped_notes)."""
    runs: List[Run] = []
    skipped: List[str] = []

    for result_file in iter_result_files(log_root):
        parts = result_file.parts
        try:
            eval_idx = parts.index("eval")
            rmbench_idx = parts.index("rmbench", eval_idx)
        except ValueError:
            continue
        if eval_idx == 0:
            continue

        exp = parts[eval_idx - 1]
        exp_task = extract_task_from_exp(exp)
        if exp_task is None:
            continue

        try:
            task = parts[rmbench_idx + 1]
            split = parts[rmbench_idx + 2]
            after_split = parts[rmbench_idx + 3]
        except IndexError:
            continue

        if split != EVAL_SPLIT:
            continue

        if MODEL_DIR_RE.match(after_split):
            model = after_split
            try:
                timestamp = parts[rmbench_idx + 4]
            except IndexError:
                continue
        else:
            model = NO_MODEL
            timestamp = after_split

        if task != exp_task:
            skipped.append(
                f"skipping task mismatch: exp={exp} exp_task={exp_task} eval_task={task} "
                f"({result_file})"
            )
            continue

        parsed = parse_result(result_file)
        if parsed is None:
            continue
        episodes, success = parsed
        if episodes == 0:
            continue

        runs.append(
            Run(
                exp=exp,
                exp_task=exp_task,
                task=task,
                model=model,
                timestamp=timestamp,
                episodes=episodes,
                success=success,
                rate=success / episodes,
                path=result_file,
            )
        )

    return runs, skipped


def pct(rate: float) -> float:
    return round(rate * 100, 2)


def run_rank_key(run: Run) -> Tuple[float, Tuple[int, str], str]:
    return (-run.rate, model_sort_key(run.model), run.timestamp)


def build_text_report(
    per_exp: Dict[str, List[Run]],
    by_task: Dict[str, List[Run]],
    skipped: List[str],
) -> str:
    lines: List[str] = []
    lines.append("=" * 78)
    lines.append(
        f"RMBench no_spatial single-task valid-evaluation ranking"
        f" (valid = {VALID_EPISODES} episodes, split={EVAL_SPLIT})"
    )
    lines.append("=" * 78)

    lines.append("")
    lines.append("#" * 78)
    lines.append("# 1. Within-experiment ranking (no_spatial single-task experiments)")
    lines.append("#" * 78)
    if not per_exp:
        lines.append("(no valid evaluation found)")
    for exp in sorted(per_exp):
        exp_runs = sorted(per_exp[exp], key=run_rank_key)
        task = exp_runs[0].exp_task if exp_runs else "?"
        lines.append("")
        lines.append(
            f"=== experiment: {exp}  task={task}  ({len(exp_runs)} valid evaluations) ==="
        )
        lines.append(
            f"{'rank':<6}{'success':>9}  {'model':<14}{'timestamp':<44}{'success/total'}"
        )
        for i, r in enumerate(exp_runs, 1):
            lines.append(
                f"{i:<4}{pct(r.rate):>7.2f}%  {r.model:<14}{r.timestamp:<44}"
                f"{r.success}/{r.episodes}"
            )

    lines.append("")
    lines.append("#" * 78)
    lines.append("# 2. Cross-experiment task ranking (no_spatial, grouped by task)")
    lines.append("#" * 78)
    if not by_task:
        lines.append("(no valid evaluation found)")
    for task in sorted(by_task):
        task_runs = sorted(by_task[task], key=run_rank_key)
        lines.append("")
        lines.append(f"=== task: {task}  ({len(task_runs)} valid evaluations) ===")
        lines.append(
            f"{'rank':<6}{'success':>9}  {'experiment':<56}{'model':<14}{'timestamp'}"
        )
        for i, r in enumerate(task_runs, 1):
            lines.append(
                f"{i:<4}{pct(r.rate):>7.2f}%  {r.exp:<58}{r.model:<14}{r.timestamp}"
            )

    if skipped:
        lines.append("")
        lines.append("#" * 78)
        lines.append(f"# Appendix: skipped task-mismatch records ({len(skipped)})")
        lines.append("#" * 78)
        for note in skipped[:20]:
            lines.append(note)
        if len(skipped) > 20:
            lines.append(f"... {len(skipped) - 20} more not shown")

    lines.append("")
    return "\n".join(lines)


THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
GROUP_FILL = PatternFill("solid", fgColor="FCE4D6")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def _style_header(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        c = ws.cell(row, col)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = CENTER


def write_per_exp_sheet(ws, per_exp: Dict[str, List[Run]]) -> None:
    headers = ["experiment", "task", "rank", "success (%)", "model", "timestamp", "success/total"]
    ncol = len(headers)
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header(ws, 1, ncol)

    row = 2
    for exp in sorted(per_exp):
        exp_runs = sorted(per_exp[exp], key=run_rank_key)
        task = exp_runs[0].exp_task if exp_runs else ""
        for i, r in enumerate(exp_runs, 1):
            values = [
                exp if i == 1 else "",
                task if i == 1 else "",
                i,
                pct(r.rate),
                r.model,
                r.timestamp,
                f"{r.success}/{r.episodes}",
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row, col, v)
                c.border = BORDER
                c.alignment = LEFT if col in (1, 2, 6) else CENTER
                if col == 4:
                    c.number_format = "0.00"
                if col == 1 and v:
                    c.font = Font(bold=True)
                    c.fill = GROUP_FILL
            row += 1

    _set_widths(ws, [44, 22, 6, 11, 14, 44, 12])


def write_by_task_sheet(ws, by_task: Dict[str, List[Run]]) -> None:
    headers = ["task", "rank", "success (%)", "experiment", "model", "timestamp", "success/total"]
    ncol = len(headers)
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    _style_header(ws, 1, ncol)

    row = 2
    for task in sorted(by_task):
        task_runs = sorted(by_task[task], key=run_rank_key)
        for i, r in enumerate(task_runs, 1):
            values = [
                task if i == 1 else "",
                i,
                pct(r.rate),
                r.exp,
                r.model,
                r.timestamp,
                f"{r.success}/{r.episodes}",
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row, col, v)
                c.border = BORDER
                c.alignment = LEFT if col in (1, 4, 6) else CENTER
                if col == 3:
                    c.number_format = "0.00"
                if col == 1 and v:
                    c.font = Font(bold=True)
                    c.fill = GROUP_FILL
            row += 1

    _set_widths(ws, [22, 6, 11, 44, 14, 44, 12])


def _set_widths(ws, widths: List[int]) -> None:
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_excel(
    per_exp: Dict[str, List[Run]],
    by_task: Dict[str, List[Run]],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "within-experiment"
    write_per_exp_sheet(ws1, per_exp)

    ws2 = wb.create_sheet("cross-experiment by task")
    write_by_task_sheet(ws2, by_task)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    log_root = LOG_ROOT.expanduser().resolve()
    if not log_root.is_dir():
        raise FileNotFoundError(f"log root does not exist: {log_root}")

    runs, skipped = collect_runs(log_root)
    valid_runs = [r for r in runs if r.is_valid]

    per_exp: Dict[str, List[Run]] = {}
    for r in valid_runs:
        per_exp.setdefault(r.exp, []).append(r)

    by_task: Dict[str, List[Run]] = {}
    for r in valid_runs:
        by_task.setdefault(r.exp_task, []).append(r)

    no_spatial_exps = sorted(
        {d.name for d in log_root.iterdir() if d.is_dir() and NO_SPATIAL_MARKER in d.name}
    )

    report = build_text_report(per_exp, by_task, skipped)
    print(report)

    OUTPUT_TXT.write_text(report, encoding="utf-8")
    write_excel(per_exp, by_task, OUTPUT_XLSX)

    incomplete = [r for r in runs if not r.is_valid]
    print("-" * 78)
    print(f"no_spatial experiment directories: {len(no_spatial_exps)}")
    for name in no_spatial_exps:
        exp_task = extract_task_from_exp(name) or "?"
        print(f"  - {name}  (task={exp_task})")
    print(
        f"Scanned {len(runs)} no_spatial results, "
        f"{len(valid_runs)} valid ({VALID_EPISODES} episodes), "
        f"{len(incomplete)} incomplete."
    )
    if incomplete:
        print("Evaluations with fewer than 100 episodes (excluded from the ranking):")
        for r in sorted(incomplete, key=lambda x: (x.exp, model_sort_key(x.model))):
            print(
                f"  {r.exp}  {r.model}  {r.timestamp}  "
                f"{r.success}/{r.episodes} ({pct(r.rate):.2f}%)"
            )
    print(f"Within-experiment ranking covers {len(per_exp)} experiments; {len(by_task)} tasks across experiments.")
    if skipped:
        print(f"Skipped {len(skipped)} task-mismatch records.")
    print(f"Wrote: {OUTPUT_TXT}")
    print(f"Wrote: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
